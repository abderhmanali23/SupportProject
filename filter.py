import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Border,Side
from openpyxl import load_workbook


def color():
    red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type="darkGrid")
    border = Border(left=Side(border_style='thin'), 
                    right=Side(border_style='thin'), 
                    top=Side(border_style='thin'), 
                    bottom=Side(border_style='thin'))
    return[red,green,border]

excel_data = pd.read_excel('support doc.xlsx', sheet_name=None)

modified_sheet_data = excel_data['Number of solved problems']
handles = [handle for handle in modified_sheet_data['Week1']]
workbook = load_workbook('support doc.xlsx')

worksheets = workbook['Number of solved problems']

column = 3
red,green,border = color()
for r_idx in range(1,len(handles)+1):
    if (len(handles[r_idx-1]) < 7) or (int(handles[r_idx-1][1:3]) <= 18):
        for col in range(1,column+1):
            worksheets.cell(row=r_idx+1, column=col).fill = red 
            worksheets.cell(row=r_idx+1, column=col).border = border

workbook.save('support doc.xlsx')