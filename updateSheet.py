import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Border,Side
from openpyxl import load_workbook

def color():
    red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type="darkGrid")
    border = Border(left=Side(border_style='thin'), 
                    right=Side(border_style='thin'), 
                    top=Side(border_style='thin'), 
                    bottom=Side(border_style='thin'))
    return[red,green,border]

def Sheet_count(file,info,numberOfWeek,numberOfProblems):
    excel_data = pd.read_excel(file, sheet_name=None)
    modified_sheet_data = excel_data['Number of solved problems']
    handles = list(map(lambda x: x.lower(),modified_sheet_data['Codeforces handle']))
    workbook = load_workbook(file)

    worksheets = workbook['Number of solved problems']
    week = []
    for handle in handles:
        try:
            week.append(f'({info[handle]}/{numberOfProblems})')
        except:
            week.append('--')

    column = numberOfWeek+ 3
    worksheets.cell(row=1, column=column,value=f'Week{numberOfWeek}')
    for r_idx in range(len(modified_sheet_data.values)):
        try:
            worksheets.cell(row=r_idx+2, column=column,value=week[r_idx])
        except:
            worksheets.cell(row=r_idx+2, column=column,value='(.../26)')

   
    workbook.save(file)

def contest_att(file,info,numberOfWeek):
    excel_data = pd.read_excel(file, sheet_name=None)
    modified_sheet_data = excel_data['Contests']
    handles = list(map(lambda x: x.lower(),modified_sheet_data['Codeforces handle']))
    workbook = load_workbook(file)

    worksheets = workbook['Contests']

    contest = []
    for handle in handles:
        contest.append(True if ("."not in handle) and (handle in info) else False)

    column = numberOfWeek + 3
    red,green,border = color()
    for r_idx, row in enumerate(modified_sheet_data.values, start=2):
        try:
            worksheets.cell(row=r_idx, column=column).fill = green if contest[r_idx-2] else red
        except:
            worksheets.cell(row=r_idx, column=column).fill = red
        worksheets.cell(row=r_idx, column=column).border = border

    workbook.save(file)
    

def session_att(file,info,numberOfWeek):
    excel_data = pd.read_excel(file, sheet_name=None)

    modified_sheet_data = excel_data['Session attendance']
    Names = [handle for handle in modified_sheet_data['Teams Name']]
    workbook = load_workbook(file)

    worksheets = workbook['Session attendance']
    contest = []
    for name in Names:
        contest.append(True if ("."not in name) and (name in info) and (info[name]) else False)

    column = numberOfWeek + 3
    red,green,border = color()
    for r_idx, row in enumerate(modified_sheet_data.values, start=2):
        try:
            worksheets.cell(row=r_idx, column=column).fill = green if contest[r_idx-1] else red
        except:
            worksheets.cell(row=r_idx, column=column).fill = red
        worksheets.cell(row=r_idx, column=column).border = border

    workbook.save(file)
    