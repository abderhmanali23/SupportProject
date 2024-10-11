import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Border,Side
from openpyxl import load_workbook

def splitNumber(string):
    return int(string[1:-1].split('/')[0])

colorRed = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
teams_data = pd.read_excel('support doc.xlsx', sheet_name=None)

modified_teams_data = teams_data['Number of solved problems']
One = [Name for Name in modified_teams_data['Week1']]
Two = [Name for Name in modified_teams_data['Week2']]
Three = [Name for Name in modified_teams_data['Week3']]

red = [False]*len(One)
for i in range(len(One)):
    if splitNumber(Three[i]) < 15:
        red[i] = True

workbook = load_workbook('support doc.xlsx')
worksheets = workbook['Number of solved problems']

for row in range(2,len(red)+2):
    if red[row-2]:
        for column in range(1,10):
            worksheets.cell(row=row,column=column).fill = colorRed

workbook.save('support doc.xlsx')